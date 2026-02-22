"""Tests for the Excel .xlsx file handler."""

import os
from pathlib import Path

import pytest
from openpyxl import load_workbook

from cowork_shield.detection.engine import DetectionEngine
from cowork_shield.exceptions import ColumnSelectionError, XLSXContentLossRiskError
from cowork_shield.handlers.xlsx import XlsxHandler
from cowork_shield.tokenizer.generator import TokenGenerator

FIXTURES_DIR = Path(__file__).parent.parent / "fixtures"


@pytest.fixture(scope="module")
def detection_engine():
    return DetectionEngine(score_threshold=0.5)


@pytest.fixture
def token_generator():
    return TokenGenerator(os.urandom(32))


class TestXlsxHandler:
    def test_anonymize_basic(self, tmp_path, detection_engine, token_generator):
        handler = XlsxHandler()
        input_path = FIXTURES_DIR / "sample_contacts.xlsx"
        output_path = tmp_path / "anonymized.xlsx"

        records, file_record = handler.anonymize(
            input_path, output_path, detection_engine, token_generator
        )

        assert output_path.exists()
        assert file_record.format == "xlsx"
        assert file_record.entities_found > 0
        assert len(records) > 0

    def test_formulas_preserved(self, tmp_path, detection_engine, token_generator):
        handler = XlsxHandler()
        input_path = FIXTURES_DIR / "sample_financial.xlsx"
        output_path = tmp_path / "anonymized.xlsx"

        handler.anonymize(input_path, output_path, detection_engine, token_generator)

        wb = load_workbook(str(output_path), data_only=False)
        ws = wb.active

        # Check that formulas are preserved
        assert ws["F2"].value == "=SUM(B2:E2)"
        assert ws["F3"].value == "=SUM(B3:E3)"
        assert ws["B5"].value == "=SUM(B2:B4)"

        wb.close()

    def test_numeric_cells_untouched(self, tmp_path, detection_engine, token_generator):
        handler = XlsxHandler()
        input_path = FIXTURES_DIR / "sample_financial.xlsx"
        output_path = tmp_path / "anonymized.xlsx"

        handler.anonymize(input_path, output_path, detection_engine, token_generator)

        wb = load_workbook(str(output_path), data_only=False)
        ws = wb.active

        # Numeric cells should be exactly preserved
        assert ws["B2"].value == 50000
        assert ws["C3"].value == 80000

        wb.close()

    def test_round_trip(self, tmp_path, detection_engine, token_generator):
        handler = XlsxHandler()
        input_path = FIXTURES_DIR / "sample_contacts.xlsx"
        anon_path = tmp_path / "anonymized.xlsx"
        restored_path = tmp_path / "restored.xlsx"

        handler.anonymize(input_path, anon_path, detection_engine, token_generator)

        reverse_lookup = token_generator.get_reverse_lookup()
        handler.restore(anon_path, restored_path, reverse_lookup)

        # Compare cell values
        wb_orig = load_workbook(str(input_path), data_only=False)
        wb_rest = load_workbook(str(restored_path), data_only=False)

        for sheet_name in wb_orig.sheetnames:
            ws_orig = wb_orig[sheet_name]
            ws_rest = wb_rest[sheet_name]
            for row_orig, row_rest in zip(ws_orig.iter_rows(), ws_rest.iter_rows()):
                for cell_orig, cell_rest in zip(row_orig, row_rest):
                    if cell_orig.value is not None:
                        assert cell_rest.value == cell_orig.value, (
                            f"Mismatch at {sheet_name}!{cell_orig.coordinate}: "
                            f"expected {cell_orig.value!r}, got {cell_rest.value!r}"
                        )

        wb_orig.close()
        wb_rest.close()

    def test_can_handle(self):
        assert XlsxHandler.can_handle(Path("data.xlsx"))
        assert XlsxHandler.can_handle(Path("data.XLSX"))
        assert not XlsxHandler.can_handle(Path("data.csv"))

    def test_backup_created(self, tmp_path, detection_engine, token_generator):
        handler = XlsxHandler()
        # Copy fixture to tmp so backup can be created next to it
        import shutil
        input_path = tmp_path / "contacts.xlsx"
        shutil.copy2(FIXTURES_DIR / "sample_contacts.xlsx", input_path)
        output_path = tmp_path / "contacts.anonymized.xlsx"

        handler.anonymize(input_path, output_path, detection_engine, token_generator)

        backup_path = input_path.with_suffix(".xlsx.backup")
        assert backup_path.exists()

    def test_blocks_lossy_content_without_ack(
        self, tmp_path, detection_engine, token_generator, monkeypatch
    ):
        monkeypatch.setattr(
            XlsxHandler,
            "_has_lossy_content",
            staticmethod(lambda wb: True),
        )
        handler = XlsxHandler(allow_lossy_xlsx=False)
        input_path = FIXTURES_DIR / "sample_contacts.xlsx"
        output_path = tmp_path / "anonymized.xlsx"

        with pytest.raises(XLSXContentLossRiskError):
            handler.anonymize(input_path, output_path, detection_engine, token_generator)

    def test_allows_lossy_content_with_ack(
        self, tmp_path, detection_engine, token_generator, monkeypatch
    ):
        monkeypatch.setattr(
            XlsxHandler,
            "_has_lossy_content",
            staticmethod(lambda wb: True),
        )
        handler = XlsxHandler(allow_lossy_xlsx=True)
        input_path = FIXTURES_DIR / "sample_contacts.xlsx"
        output_path = tmp_path / "anonymized.xlsx"

        handler.anonymize(input_path, output_path, detection_engine, token_generator)
        assert output_path.exists()

    def test_inspect_columns(self):
        handler = XlsxHandler()
        columns = handler.inspect_columns(FIXTURES_DIR / "sample_contacts.xlsx")
        assert len(columns) >= 4
        assert columns[0].letter == "A"
        assert columns[0].name.lower() in {"name", "full name"}
        assert columns[0].data_type == "text"
        assert columns[0].sample_values

    def test_column_mode_by_letter(self, tmp_path, detection_engine, token_generator):
        handler = XlsxHandler()
        input_path = FIXTURES_DIR / "sample_contacts.xlsx"
        output_path = tmp_path / "column_only.xlsx"

        handler.anonymize(
            input_path,
            output_path,
            detection_engine,
            token_generator,
            selected_columns=["A"],
            detect_pii=False,
        )

        wb = load_workbook(str(output_path), data_only=False)
        ws = wb.active
        assert isinstance(ws["A2"].value, str)
        assert ws["A2"].value.startswith("[COL_A_")
        # Non-selected column should remain plaintext when detect_pii is false.
        assert ws["B2"].value is not None
        wb.close()

    def test_column_mode_by_name_prefix(self, tmp_path, detection_engine, token_generator):
        handler = XlsxHandler()
        input_path = FIXTURES_DIR / "sample_contacts.xlsx"
        output_path = tmp_path / "column_name.xlsx"

        handler.anonymize(
            input_path,
            output_path,
            detection_engine,
            token_generator,
            selected_columns=["Name"],
            detect_pii=False,
        )

        wb = load_workbook(str(output_path), data_only=False)
        ws = wb.active
        assert isinstance(ws["A2"].value, str)
        assert ws["A2"].value.startswith("[NAME_")
        wb.close()

    def test_column_mode_skips_formula_cells(
        self,
        tmp_path,
        detection_engine,
        token_generator,
    ):
        handler = XlsxHandler()
        input_path = FIXTURES_DIR / "sample_financial.xlsx"
        output_path = tmp_path / "formula_column.xlsx"

        handler.anonymize(
            input_path,
            output_path,
            detection_engine,
            token_generator,
            selected_columns=["F"],
            detect_pii=False,
        )

        wb = load_workbook(str(output_path), data_only=False)
        ws = wb.active
        assert ws["F2"].value == "=SUM(B2:E2)"
        assert ws["F3"].value == "=SUM(B3:E3)"
        wb.close()

    def test_column_mode_with_detect_pii_enabled(
        self,
        tmp_path,
        detection_engine,
        token_generator,
    ):
        handler = XlsxHandler()
        input_path = FIXTURES_DIR / "sample_contacts.xlsx"
        output_path = tmp_path / "column_plus_pii.xlsx"

        handler.anonymize(
            input_path,
            output_path,
            detection_engine,
            token_generator,
            selected_columns=["A"],
            detect_pii=True,
        )

        wb = load_workbook(str(output_path), data_only=False)
        ws = wb.active
        assert isinstance(ws["A2"].value, str)
        assert ws["A2"].value.startswith("[COL_A_")
        # Email/phone columns should be tokenized by PII detection in combined mode.
        row_text = " ".join(str(ws[f"{col}2"].value or "") for col in ("B", "C", "D"))
        assert "[" in row_text
        wb.close()

    def test_invalid_column_selection_raises(
        self,
        tmp_path,
        detection_engine,
        token_generator,
    ):
        handler = XlsxHandler()
        with pytest.raises(ColumnSelectionError):
            handler.anonymize(
                FIXTURES_DIR / "sample_contacts.xlsx",
                tmp_path / "invalid.xlsx",
                detection_engine,
                token_generator,
                selected_columns=["Missing Column"],
                detect_pii=False,
            )

    def test_column_mode_round_trip(self, tmp_path, detection_engine, token_generator):
        handler = XlsxHandler()
        input_path = FIXTURES_DIR / "sample_contacts.xlsx"
        anonymized_path = tmp_path / "column_roundtrip.anonymized.xlsx"
        restored_path = tmp_path / "column_roundtrip.restored.xlsx"

        handler.anonymize(
            input_path,
            anonymized_path,
            detection_engine,
            token_generator,
            selected_columns=["A", "C"],
            detect_pii=False,
        )
        handler.restore(anonymized_path, restored_path, token_generator.get_reverse_lookup())

        wb_orig = load_workbook(str(input_path), data_only=False)
        wb_rest = load_workbook(str(restored_path), data_only=False)
        for sheet_name in wb_orig.sheetnames:
            ws_orig = wb_orig[sheet_name]
            ws_rest = wb_rest[sheet_name]
            for row_orig, row_rest in zip(ws_orig.iter_rows(), ws_rest.iter_rows()):
                for cell_orig, cell_rest in zip(row_orig, row_rest):
                    if cell_orig.value is not None:
                        assert cell_rest.value == cell_orig.value
        wb_orig.close()
        wb_rest.close()
