"""Excel .xlsx file handler using openpyxl."""

from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from cowork_shield.detection.engine import DetectionEngine
from cowork_shield.exceptions import XLSXContentLossRiskError
from cowork_shield.handlers.column_select import (
    ColumnDescriptor,
    ColumnSelection,
    describe_columns,
    infer_data_type,
    resolve_column_selections,
)
from cowork_shield.models import (
    DetectedEntity,
    EntityType,
    FileRecord,
    ReplacementRecord,
    now_iso,
)
from cowork_shield.tokenizer.generator import TokenGenerator
from cowork_shield.tokenizer.replacer import TextReplacer
from cowork_shield.verification.verifier import compute_sha256


class XlsxHandler:
    """Handles .xlsx files with cell-level anonymization.

    Design decisions:
    - Load with data_only=False to preserve formulas
    - Skip formula cells (starting with '=') — never modify formulas
    - Skip pure numeric cells (int/float) — no PII in numbers
    - Cell formatting is preserved automatically by openpyxl

    WARNING: openpyxl can destroy charts, images, and shapes on save.
    This handler blocks by default when those features are present unless
    explicitly overridden.
    """

    def __init__(self, allow_lossy_xlsx: bool = False):
        self._replacer = TextReplacer()
        self._allow_lossy_xlsx = allow_lossy_xlsx

    @staticmethod
    def can_handle(file_path: Path) -> bool:
        return file_path.suffix.lower() == ".xlsx"

    def anonymize(
        self,
        input_path: Path,
        output_path: Path,
        detection_engine: DetectionEngine,
        token_generator: TokenGenerator,
        source_file: str = "",
        language: str = "auto",
        selected_columns: list[str] | None = None,
        detect_pii: bool = True,
    ) -> tuple[list[ReplacementRecord], FileRecord]:
        # Create backup before any modification
        backup_path = input_path.with_suffix(input_path.suffix + ".backup")
        if not backup_path.exists():
            shutil.copy2(input_path, backup_path)

        wb = load_workbook(str(input_path), data_only=False)
        if self._has_lossy_content(wb) and not self._allow_lossy_xlsx:
            wb.close()
            raise XLSXContentLossRiskError(
                "XLSX contains charts/images that openpyxl may drop. "
                "Re-run with --allow-lossy-xlsx to acknowledge risk."
            )
        all_records: list[ReplacementRecord] = []
        total_entities = 0
        baseline_selection: dict[int, ColumnSelection] = {}
        if selected_columns:
            baseline_ws = wb[wb.sheetnames[0]]
            baseline_selection = self._resolve_sheet_column_selection(
                baseline_ws,
                selected_columns,
            )

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            selected_map = self._project_selection_to_sheet(
                ws,
                baseline_selection,
            )
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue

                    # Skip formula cells
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        continue

                    value_str = str(cell.value)
                    if not value_str.strip():
                        continue

                    selection = selected_map.get(cell.column - 1)
                    if selection is not None:
                        # Preserve header row names when selecting by column.
                        if cell.row == 1:
                            continue
                        source_id = (
                            f"{sheet_name}!{get_column_letter(cell.column)}{cell.row}"
                        )
                        token = token_generator.get_or_create_column_token(
                            value_str,
                            selection.token_prefix,
                            source_file=source_file,
                        )
                        cell.value = token.token_text
                        all_records.append(
                            ReplacementRecord(
                                location=source_id,
                                original_value=value_str,
                                token_text=token.token_text,
                                entity_type=EntityType.COLUMN,
                            )
                        )
                        total_entities += 1
                        continue

                    if not detect_pii:
                        continue

                    # Skip pure numeric cells for detection mode.
                    if isinstance(cell.value, (int, float)):
                        continue

                    source_id = (
                        f"{sheet_name}!{get_column_letter(cell.column)}{cell.row}"
                    )

                    entities = _detect_entities(
                        detection_engine,
                        text=value_str,
                        source_id=source_id,
                        language=language,
                    )
                    total_entities += len(entities)

                    if entities:
                        replaced, records = self._replacer.replace_entities(
                            value_str, entities, token_generator, source_file
                        )
                        cell.value = replaced
                        all_records.extend(records)

        wb.save(str(output_path))
        wb.close()

        hash_before = compute_sha256(input_path)
        hash_after = compute_sha256(output_path)

        file_record = FileRecord(
            file_path=str(input_path),
            file_hash_before=hash_before,
            file_hash_after=hash_after,
            anonymized_path=str(output_path),
            entities_found=total_entities,
            tokens_applied=len(all_records),
            timestamp=now_iso(),
            format="xlsx",
            applied_tokens=sorted({record.token_text for record in all_records}),
        )

        return all_records, file_record

    def inspect_columns(self, input_path: Path) -> list[ColumnDescriptor]:
        """Return available columns from the first worksheet."""
        wb = load_workbook(str(input_path), data_only=False, read_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            max_columns = ws.max_column or 0
            headers = self._sheet_headers(ws, max_columns)
            samples = self._sheet_samples(ws, max_columns)
            data_types = {
                idx: infer_data_type(values)
                for idx, values in samples.items()
            }
            return describe_columns(
                headers,
                max_columns,
                sample_values=samples,
                data_types=data_types,
            )
        finally:
            wb.close()

    def restore(
        self,
        input_path: Path,
        output_path: Path,
        reverse_lookup: dict[str, str],
    ) -> None:
        wb = load_workbook(str(input_path), data_only=False)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue

                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        continue

                    value_str = str(cell.value)
                    restored = self._replacer.restore_tokens(value_str, reverse_lookup)
                    if restored != value_str:
                        cell.value = restored

        wb.save(str(output_path))
        wb.close()

    @staticmethod
    def _has_lossy_content(workbook) -> bool:
        for worksheet in workbook.worksheets:
            if getattr(worksheet, "_charts", None):
                if len(worksheet._charts) > 0:
                    return True
            if getattr(worksheet, "_images", None):
                if len(worksheet._images) > 0:
                    return True
        return False

    @staticmethod
    def _sheet_headers(worksheet, max_columns: int) -> list[str]:
        headers: list[str] = []
        for idx in range(1, max_columns + 1):
            value = worksheet.cell(row=1, column=idx).value
            headers.append(str(value).strip() if value is not None else "")
        return headers

    @staticmethod
    def _sheet_samples(worksheet, max_columns: int) -> dict[int, list[str]]:
        samples: dict[int, list[str]] = {idx: [] for idx in range(max_columns)}
        max_rows = worksheet.max_row or 0
        for row_idx in range(2, max_rows + 1):
            for idx in range(max_columns):
                if len(samples[idx]) >= 3:
                    continue
                value = worksheet.cell(row=row_idx, column=idx + 1).value
                if value is None:
                    continue
                value_str = str(value).strip()
                if value_str:
                    samples[idx].append(value_str[:40])
            if all(len(values) >= 3 for values in samples.values()):
                break
        return samples

    def _resolve_sheet_column_selection(
        self,
        worksheet,
        selected_columns: list[str],
    ) -> dict[int, ColumnSelection]:
        max_columns = worksheet.max_column or 0
        headers = self._sheet_headers(worksheet, max_columns)
        return resolve_column_selections(selected_columns, headers=headers, max_columns=max_columns)

    @staticmethod
    def _project_selection_to_sheet(
        worksheet,
        baseline_selection: dict[int, ColumnSelection],
    ) -> dict[int, ColumnSelection]:
        if not baseline_selection:
            return {}
        max_columns = worksheet.max_column or 0
        return {
            index: selection
            for index, selection in baseline_selection.items()
            if index < max_columns
        }


def _detect_entities(
    detection_engine: DetectionEngine,
    *,
    text: str,
    source_id: str,
    language: str,
) -> list[DetectedEntity]:
    try:
        return detection_engine.detect_in_cell(text, source_id, language=language)
    except TypeError:
        # Compatibility for tests using stub engines without language arg.
        return detection_engine.detect_in_cell(text, source_id)
